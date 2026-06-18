#!/usr/bin/env python3
# Exporta las facturas de caja chica para el contador:
#  - Excel con pestañas por empresa (Super Techos / Prouco) + Resumen + Revisar
#  - Valida NCF y RNC (marca faltantes y errores numéricos)
#  - Miniatura embebida + link a la foto original (carpeta fotos/ en el mismo ZIP)
# Uso: python3 export-caja-chica-contador.py [YYYY-MM]
#   sin argumento = todas; con "2026-05" = solo ese mes.
import os, re, io, sys, json, zipfile, urllib.request, urllib.parse, ssl
from datetime import date
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
HOME = os.path.expanduser("~")
MESES = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
         "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
# Filtro de mes opcional (YYYY-MM)
MES = sys.argv[1] if len(sys.argv) > 1 else None
if MES:
    y, m = int(MES[:4]), int(MES[5:7])
    DESDE = date(y, m, 1).isoformat()
    HASTA = (date(y + (m == 12), (m % 12) + 1, 1)).isoformat()
    SUF = f"{MESES[m]}-{y}"
else:
    DESDE = HASTA = None
    SUF = "Todas"
NOMBRE = f"Caja-Chica-{SUF}"
OUTDIR = os.path.join(HOME, "Downloads", NOMBRE)
FOTOS = os.path.join(OUTDIR, "fotos")
THUMBS = "/tmp/cc_thumbs"
XLSX = os.path.join(OUTDIR, f"{NOMBRE}.xlsx")
ZIPOUT = os.path.join(HOME, "Downloads", f"{NOMBRE}.zip")
BUCKET = "caja-chica-facturas"
for d in (OUTDIR, FOTOS, THUMBS):
    os.makedirs(d, exist_ok=True)

# ---- credenciales: usa .env.prod (producción) si existe, si no .env.local ----
def cargar_env(path):
    d = {}
    if not os.path.exists(path): return d
    for ln in open(path):
        ln = ln.strip()
        if ln and not ln.startswith("#") and "=" in ln:
            k, v = ln.split("=", 1)
            d[k.strip()] = v.strip().strip('"').strip("'")
    return d

env = cargar_env(os.path.join(REPO, ".env.prod")) or cargar_env(os.path.join(REPO, ".env.local"))
URL = env.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
SVC = env.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
ANON = env.get("NEXT_PUBLIC_SUPABASE_KEY", "").strip()
DATA_KEY = ANON or SVC  # los datos se leen con la pública (RLS off)
if "127.0.0.1" in URL or "localhost" in URL or not DATA_KEY:
    raise SystemExit("Falta .env.prod con la URL y NEXT_PUBLIC_SUPABASE_KEY de PRODUCCIÓN.")
HAY_FOTOS = bool(SVC)  # el bucket es privado: solo con service_role se bajan fotos
CTX = ssl.create_default_context()

def api_get(path):
    req = urllib.request.Request(URL + path, headers={
        "apikey": DATA_KEY, "Authorization": "Bearer " + DATA_KEY, "Accept": "application/json"})
    with urllib.request.urlopen(req, context=CTX, timeout=60) as r:
        return json.loads(r.read().decode())

def storage_get(p):
    req = urllib.request.Request(
        URL + "/storage/v1/object/" + BUCKET + "/" + urllib.parse.quote(p),
        headers={"apikey": SVC, "Authorization": "Bearer " + SVC})
    with urllib.request.urlopen(req, context=CTX, timeout=60) as r:
        return r.read()

# ---- datos ----
print("Descargando datos...")
cols = "id,fecha,proveedor,rnc,monto,concepto,foto_path,rotacion_grados,empresa_receptora,status,created_at,creado_por_id,datos_ia"
filtro = "&tipo=eq.gasto_factura"
if DESDE: filtro += f"&fecha=gte.{DESDE}&fecha=lt.{HASTA}"
rows = api_get(f"/rest/v1/caja_chica_movimientos?select={cols}{filtro}&order=fecha.desc&limit=5000")
personal = {p["id"]: p["nombre"] for p in api_get("/rest/v1/personal?select=id,nombre")}
print(f"  {len(rows)} facturas")

# ---- validaciones ----
def val_rnc(raw):
    # Consistente con la app (lib/validacionFiscal.js): la CANTIDAD de dígitos es lo
    # "duro"; el dígito verificador DGII solo se marca para verificar (35% de falsos
    # positivos contra RNC reales), no como error que va a la pestaña Revisar.
    s = re.sub(r"\D", "", raw or "")
    if not s: return ("Falta", False)
    if len(s) == 11: return ("Cédula", True)
    if len(s) != 9: return ("Longitud inválida", False)
    w = [7, 9, 8, 6, 5, 4, 3, 2]
    t = sum(int(s[i]) * w[i] for i in range(8)) % 11
    dv = 2 if t == 0 else (1 if t == 1 else 11 - t)
    return ("OK", True) if dv == int(s[8]) else ("Verificar dígito", True)

def val_ncf(raw):
    s = (raw or "").strip().upper()
    if not s: return ("Falta", False)
    if re.match(r"^[A-Z]\d{10}$", s): return ("OK", True)
    if re.match(r"^E\d{12}$", s): return ("OK (e-CF)", True)
    return ("Revisar formato", False)

EMPRESA_LABEL = {"super_techos": "Super Techos", "prouco": "Prouco"}
def empresa_de(r):
    e = r.get("empresa_receptora") or (r.get("datos_ia") or {}).get("empresa_receptora") or ""
    return EMPRESA_LABEL.get(e, "Sin empresa")

# ---- fotos ----
fotos_ok = {}
if not HAY_FOTOS:
    print("Sin llave service_role: se genera el Excel SOLO con datos (sin fotos ni links).")
else:
  print("Descargando fotos...")
  for r in rows:
    fp = r.get("foto_path")
    if not fp: continue
    fname = os.path.basename(fp)
    dest = os.path.join(FOTOS, fname)
    try:
        if not os.path.exists(dest):
            data = storage_get(fp)
            im = Image.open(io.BytesIO(data)).convert("RGB")
            deg = r.get("rotacion_grados") or 0
            if deg: im = im.rotate(-deg, expand=True)
            im.save(dest, "JPEG", quality=85)
        else:
            im = Image.open(dest)
        th = im.copy(); th.thumbnail((110, 110))
        tp = os.path.join(THUMBS, fname)
        th.save(tp, "JPEG", quality=80)
        fotos_ok[r["id"]] = (fname, tp)
    except Exception as e:
        print("  ! foto", fp, e)
print(f"  {len(fotos_ok)} fotos descargadas")

# ---- construir registros ----
recs = []
for r in rows:
    di = r.get("datos_ia") or {}
    ncf = di.get("ncf") or ""
    rnc = r.get("rnc") or di.get("rnc") or ""
    rnc_st, rnc_ok = val_rnc(rnc)
    ncf_st, ncf_ok = val_ncf(ncf)
    recs.append({
        "empresa": empresa_de(r),
        "fecha": r.get("fecha") or "",
        "proveedor": r.get("proveedor") or "",
        "rnc": rnc, "rnc_st": rnc_st, "rnc_ok": rnc_ok,
        "ncf": ncf, "ncf_st": ncf_st, "ncf_ok": ncf_ok,
        "concepto": r.get("concepto") or "",
        "categoria": di.get("categoria_sugerida") or "",
        "subtotal": di.get("subtotal"),
        "itbis": di.get("itbis"),
        "monto": float(r.get("monto") or 0),
        "estado": r.get("status") or "",
        "registrado": personal.get(r.get("creado_por_id"), ""),
        "foto": fotos_ok.get(r["id"]),
    })

# ---- estilos ----
DARK = "FF0E0E11"; HEADW = Font(bold=True, color="FFFFFFFF", size=10, name="Arial")
RED = PatternFill("solid", fgColor="FFF8DADA"); AMBER = PatternFill("solid", fgColor="FFFCEEDA")
GREEN = PatternFill("solid", fgColor="FFE7F3E1"); HFILL = PatternFill("solid", fgColor=DARK)
BORD = Border(bottom=Side(style="thin", color="FFE2E2E6"))
LINKF = Font(color="FF185FA5", underline="single", name="Arial", size=10)
CELLF = Font(name="Arial", size=10)
COLS = [("Fecha", 11), ("Proveedor", 26), ("RNC", 13), ("RNC ✓", 14), ("NCF", 14), ("NCF ✓", 14),
        ("Concepto", 30), ("Categoría", 14), ("Subtotal", 12), ("ITBIS", 11), ("Monto RD$", 13),
        ("Estado", 11), ("Registró", 18), ("Foto", 17), ("Link", 9)]

def money(v):
    try: return float(v)
    except: return None

def hoja_facturas(ws, registros):
    ws.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(1, i, h); c.font = HEADW; c.fill = HFILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    rr = 2
    for d in registros:
        ws.row_dimensions[rr].height = 78 if d["foto"] else 18
        vals = [d["fecha"], d["proveedor"], d["rnc"], d["rnc_st"], d["ncf"], d["ncf_st"],
                d["concepto"], d["categoria"], money(d["subtotal"]), money(d["itbis"]),
                d["monto"], d["estado"], d["registrado"], "", ""]
        for i, v in enumerate(vals, 1):
            c = ws.cell(rr, i, v); c.font = CELLF; c.border = BORD
            c.alignment = Alignment(vertical="center", wrap_text=(i in (2, 7)))
            if i in (9, 10, 11) and isinstance(v, float): c.number_format = '#,##0.00'
        # validaciones coloreadas
        if not d["rnc_ok"]: ws.cell(rr, 4).fill = (AMBER if d["rnc_st"] == "Falta" else RED)
        elif d["rnc_st"] == "Verificar dígito": ws.cell(rr, 4).fill = AMBER
        else: ws.cell(rr, 4).fill = GREEN
        if not d["ncf_ok"]: ws.cell(rr, 6).fill = (AMBER if d["ncf_st"] == "Falta" else RED)
        else: ws.cell(rr, 6).fill = GREEN
        # foto + link
        if d["foto"]:
            fname, tp = d["foto"]
            img = XLImage(tp); ws.add_image(img, f"N{rr}")
            lc = ws.cell(rr, 15, "Abrir"); lc.hyperlink = f"fotos/{fname}"; lc.font = LINKF
            lc.alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws.cell(rr, 14, "—").alignment = Alignment(horizontal="center")
        rr += 1
    ws.auto_filter.ref = f"A1:O{rr-1}"

wb = openpyxl.Workbook()
# Resumen primero
res = wb.active; res.title = "Resumen"; res.sheet_view.showGridLines = False

por_emp = {}
for d in recs:
    por_emp.setdefault(d["empresa"], []).append(d)

# Hojas por empresa (orden fijo)
orden = ["Super Techos", "Prouco"] + [e for e in por_emp if e not in ("Super Techos", "Prouco")]
for emp in orden:
    if emp not in por_emp: continue
    ws = wb.create_sheet(emp[:31])
    hoja_facturas(ws, por_emp[emp])

# Hoja Revisar (solo errores numéricos de NCF/RNC)
problemas = [d for d in recs if d["ncf_st"] == "Revisar formato" or d["rnc_st"] == "Longitud inválida"]
wr = wb.create_sheet("Revisar"); wr.sheet_view.showGridLines = False
hr = [("Empresa", 14), ("Fecha", 11), ("Proveedor", 26), ("RNC", 13), ("RNC ✓", 16),
      ("NCF", 14), ("NCF ✓", 16), ("Monto RD$", 13), ("Link", 9)]
for i, (h, w) in enumerate(hr, 1):
    wr.column_dimensions[get_column_letter(i)].width = w
    c = wr.cell(1, i, h); c.font = HEADW; c.fill = HFILL
    c.alignment = Alignment(horizontal="center")
wr.freeze_panes = "A2"; wr.row_dimensions[1].height = 22
for j, d in enumerate(problemas, 2):
    row = [d["empresa"], d["fecha"], d["proveedor"], d["rnc"], d["rnc_st"], d["ncf"], d["ncf_st"], d["monto"], ""]
    for i, v in enumerate(row, 1):
        c = wr.cell(j, i, v); c.font = CELLF; c.border = BORD
        if i == 8 and isinstance(v, float): c.number_format = '#,##0.00'
    if not d["rnc_ok"] and d["rnc_st"] != "Falta": wr.cell(j, 5).fill = RED
    if not d["ncf_ok"] and d["ncf_st"] != "Falta": wr.cell(j, 7).fill = RED
    if d["foto"]:
        lc = wr.cell(j, 9, "Abrir"); lc.hyperlink = f"fotos/{d['foto'][0]}"; lc.font = LINKF
wr.auto_filter.ref = f"A1:I{len(problemas)+1}"

# Rellenar Resumen
res.column_dimensions["A"].width = 26
for i, w in zip("BCDEFG", [13, 16, 12, 12, 12, 12]): res.column_dimensions[i].width = w
res["A1"] = "Caja chica — Facturas para contabilidad"; res["A1"].font = Font(bold=True, size=14, name="Arial")
res["A2"] = "Generado para el contador · datos reales del ERP (tipo gasto_factura)"; res["A2"].font = Font(size=9, color="FF6B6B70", name="Arial")
hdr = ["Empresa", "Facturas", "Monto total", "Con foto", "Sin NCF", "NCF inválido", "Sin/err RNC"]
for i, h in enumerate(hdr, 1):
    c = res.cell(4, i, h); c.font = HEADW; c.fill = HFILL; c.alignment = Alignment(horizontal="center")
rr = 5
for emp in orden:
    if emp not in por_emp: continue
    ds = por_emp[emp]
    vals = [emp, len(ds), sum(d["monto"] for d in ds),
            sum(1 for d in ds if d["foto"]),
            sum(1 for d in ds if d["ncf_st"] == "Falta"),
            sum(1 for d in ds if d["ncf_st"] == "Revisar formato"),
            sum(1 for d in ds if not d["rnc_ok"])]
    for i, v in enumerate(vals, 1):
        c = res.cell(rr, i, v); c.font = CELLF
        if i == 3: c.number_format = '#,##0.00'
    rr += 1
# total
tot = ["TOTAL", len(recs), sum(d["monto"] for d in recs), sum(1 for d in recs if d["foto"]),
       sum(1 for d in recs if d["ncf_st"] == "Falta"), sum(1 for d in recs if d["ncf_st"] == "Revisar formato"),
       sum(1 for d in recs if not d["rnc_ok"])]
for i, v in enumerate(tot, 1):
    c = res.cell(rr, i, v); c.font = Font(bold=True, name="Arial", size=10)
    if i == 3: c.number_format = '#,##0.00'
res.cell(rr + 2, 1, "Validaciones:").font = Font(bold=True, name="Arial", size=10)
notas = [
    "RNC: válido = 9 dígitos con dígito verificador correcto (DGII) o 11 (cédula). Rojo = error numérico; amarillo = falta.",
    "NCF: válido = letra + 10 dígitos (o e-CF). Rojo = formato a revisar; amarillo = falta.",
    "Pestaña 'Revisar' lista solo las facturas con NCF/RNC mal escritos (errores numéricos) que el contador debe corregir.",
    "Cada factura con comprobante tiene su foto embebida y un link 'Abrir' a la imagen original en la carpeta fotos/.",
]
for k, n in enumerate(notas):
    res.cell(rr + 3 + k, 1, "• " + n).font = Font(size=9, color="FF44444A", name="Arial")

wb.save(XLSX)
print("Excel:", XLSX)

# ---- ZIP ----
with zipfile.ZipFile(ZIPOUT, "w", zipfile.ZIP_DEFLATED) as z:
    z.write(XLSX, "Caja-Chica-Facturas.xlsx")
    for fn in sorted(os.listdir(FOTOS)):
        z.write(os.path.join(FOTOS, fn), f"fotos/{fn}")
print("ZIP:", ZIPOUT, "·", round(os.path.getsize(ZIPOUT) / 1024 / 1024, 1), "MB")
print("Facturas:", len(recs), "| Con foto:", len(fotos_ok), "| Revisar:", len(problemas))
